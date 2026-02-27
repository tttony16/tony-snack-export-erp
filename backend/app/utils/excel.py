from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def create_workbook(title: str, headers: list[str], rows: list[list[Any]]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, len(rows) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max_len + 4, 50
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def read_workbook(content: bytes) -> list[dict]:
    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        item = {}
        for i, header in enumerate(headers):
            item[header] = row[i] if i < len(row) else None
        result.append(item)

    wb.close()
    return result


def create_template(title: str, headers: list[str]) -> BytesIO:
    return create_workbook(title, headers, [])
